#from __future__ import division

import scipy
import numpy
import pandas
import openpyxl
import csv
import glob
import os
import random

from datetime import  datetime
import dateutil.parser as parser
from datetime import timedelta

from keras.utils import np_utils
from sklearn.preprocessing import LabelEncoder

import MathFunctions

#Importing all .csv files
def ExtractData():
    D1_train = numpy.loadtxt("WBB_SaturdayData_train.csv", delimiter=',')
    D1_test = numpy.loadtxt("WBB_SaturdayData_test_2030.csv", delimiter=',')
    D2_train = numpy.loadtxt("WBB_SundayData_train.csv", delimiter=',')
    D2_test = numpy.loadtxt("WBB_SundayData_test_2030.csv", delimiter=',')
    D3_train = numpy.loadtxt("WBB_WeekdayData_train.csv", delimiter=',')
    D3_test = numpy.loadtxt("WBB_WeekdayData_test_2030.csv", delimiter=',')

    # Creating class that can be used to create a list of objects
    class DaySelect(object):
        """__init__() functions as the class constructor"""

        def __init__(self, DataTrain=None, DataTest=None):
            self.DataTrain = DataTrain
            self.DataTest = DataTest

    print

    # Seggregating by cluster, where each cluster represents, Saturday, Sunday and Weekdays

    ClusterData = []
    ClusterData.append(DaySelect(D1_train, D1_test))
    ClusterData.append(DaySelect(D2_train, D2_test))
    ClusterData.append(DaySelect(D3_train, D3_test))

    return ClusterData

#Function to get weatherfile

def get_weatherdata(weather_file):
    wb = openpyxl.load_workbook(weather_file)
    ws = wb['Sheet1']

    T_db = numpy.array([[cell.value for cell in col] for col in ws['H9':'H8768']])
    DNI = numpy.array([[cell.value for cell in col] for col in ws['P9':'P8768']])
    RH = numpy.array([[cell.value for cell in col] for col in ws['J9':'J8768']])
    WS = numpy.array([[cell.value for cell in col] for col in ws['W9':'W8768']])

    weather_out = numpy.concatenate((T_db, DNI, RH, WS), axis=1)
    weather_max = numpy.amax(weather_out, axis=0)
    weather_min = numpy.amin(weather_out, axis=0)

    return weather_out, weather_max, weather_min

#Function to
#Function to get output file

def get_energydata(output_file):
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Sheet1']
    H = numpy.array([[cell.value for cell in col] for col in ws['B2':'B8761']])/1E9
    #H = numpy.array([[cell.value for cell in col] for col in ws['J2':'J8761']]) / 1E9
    return H


def normalize_vector(weather_in, weather_min, weather_max):
    row_max, col_max = weather_in.shape

    weather_out = numpy.zeros((row_max, col_max))

    for j in range(0, col_max):
        max_val = weather_max[j]
        min_val = weather_min[j]

        weather_out[:, j] = (weather_in[:, j] - min_val)/(max_val - min_val)

    return weather_out


# Function to create schedule
def CreateSchedule(nn_model):

    class ScheduleData(object):
        """__init__() functions as the class constructor"""

        def __init__(self, model=None):
            self.model = model

    print

    schedule = []
    schedule.append(ScheduleData(nn_model))

    return schedule


# Function to round off a real number to 0 or 1



def ConvertToBinary(D):
    row, col = D.shape
    for j in range(0, col):
        for i in range(0, row):
            if D[i, j] < 0.5:
                D[i, j] = 0
            else:
                D[i, j] = 1

    return D


def FixRealValue(D):
    row, col = D.shape
    for j in range(0, col):
        for i in range(0, row):
            if D[i, j] < 0:
                D[i, j] = 0


    return D

def find_day_idx(day):

    j1 = (day-1)*24
    j2 = day*24 - 1
    return j1, j2


def separate_data_weekday(input_train, output_train, input_test, output_test):
    HolidayList = numpy.array([315, 359, 192, 16, 51, 149, 247, 327])
    DayCount = numpy.array([51, 61, 253])

    data_size, dim = input_train.shape

    class separate_by_day(object):
        """__init__() functions as the class constructor"""

        #Initializing Matrix for each schedule
        def __init__(self, day=None, day_count=None):
            self.ID = day
            self.hours = numpy.arange(1, 25, 1)

            self.X_train = numpy.zeros((24*day_count, dim))
            self.Y_train = numpy.zeros((24*day_count, 1))
            self.X_test = numpy.zeros((24*day_count, dim))
            self.Y_test = numpy.zeros((24*day_count, 1))
            self.idx = numpy.squeeze(numpy.zeros((24*day_count, 1)))

    print

    ClusterData = []

    for day in range(0, len(DayCount)):

        ClusterData.append(separate_by_day(day, DayCount[day]))



    for day in range(0, len(DayCount)):

        ClusterData[day].counter = 0

        num = numpy.int32(len(input_train)/24)


        for i in range(0, num):
            res = (i) % 7
            z = numpy.sum(i==(HolidayList-1))

            if(z>0 or res==0):
                a = 1
            elif(res==6):
                a = 0
            else:
                a = 2



            if a==day:
                ClusterData[day].counter = ClusterData[day].counter + 1
                count = ClusterData[day].counter

                ClusterData[day].X_train[((count-1)*24):((count-1)*24 + 24), :] = input_train[(i*24):(i*24 + 24), :] #CHECK i here
                ClusterData[day].Y_train[(count - 1) * 24:(count - 1) * 24 + 24, :] = output_train[(i * 24):(i * 24 + 24), :]

                ClusterData[day].X_test[(count - 1) * 24:(count - 1) * 24 + 24, :] = input_test[(i * 24):(i * 24 + 24), :]
                ClusterData[day].Y_test[(count - 1) * 24:(count - 1) * 24 + 24] = output_test[(i* 24):(i * 24 + 24)]
                ClusterData[day].idx[(count - 1) * 24:(count - 1) * 24 + 24] = numpy.arange(i*24, (i*24 + 24))

            row, col = ClusterData[day].X_train.shape

            t24 = numpy.arange(1, 25)
            ClusterData[day].t = numpy.tile(t24, (row / 24))

    return ClusterData



###csv reader and writers
def read_csvfile(filename):
    with open(filename, 'rb') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        for row in spamreader:
            print ",".join(row)


    return spamreader


def write_csvfile(filename, arr_write):
    with open(filename, 'a') as csvfile:
        spamwriter = csv.writer(csvfile, delimiter=' ', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        spamwriter.writerow(arr_write)


def reset_csvfile(filename):
    f = open(filename, "w+")
    f.close()


def create_csvfile(choice, filename):

    if choice==1:
        reset_csvfile(filename)
        with open(filename, 'wb') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=' ', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            spamwriter.writerow(['input_dim'] + ['real_units'])




def read_PSB_csv(date_start, date_end, col_idx1, col_idx2):
    file_ = r'/home/sseslab/Documents/SLC PSB data/SLC PSB/PSB_Meters4.csv'

    #df = pandas.read_csv(file_, delimiter=',', skiprows=1, header=None)
    df = pandas.read_csv(file_, delimiter=',')
    df = df[df.DateTime != '12:00:00 AM']
    datetime_pandas = df.DateTime
    datetime_mat = datetime_pandas.as_matrix()
    datetime_mat = datetime_mat[::-1]

    start_idx, end_idx = PSB_select_date(datetime_mat, date_start, date_end)
    datetime_select = datetime_mat[start_idx:end_idx+1]
    delta, datetime_new = PSB_time_array(datetime_select)



    time_array = convert_to_timetep(delta)

    timestep_total = find_timesteps(date_start, date_end)


    df = df.apply(pandas.to_numeric, args=('coerce',))
    data_arr = df.as_matrix(columns=df.columns[col_idx1:col_idx2])
    data_arr = data_arr[::-1]

    data_arr = data_arr[start_idx:end_idx+1, :]
    print data_arr[:, 4]
    print delta.shape
    print delta
    print time_array
    #print timestep_total
    #data_arr = interpolate_nans(data_arr)


    data_arr = arrange_array(data_arr, time_array, timestep_total)
    #zero_idx = numpy.where(~data_arr.any(axis=1))[0]

    #for i in range(0, len(zero_idx)):
    #    data_arr[zero_idx[i], :] = numpy.nan


    #data_arr = interpolate_nans(data_arr)

    return data_arr


def read_muliple_csv(col_idx1, col_idx2):
    folder_path = r'/home/sseslab/Documents/SLC PSB data/SLC PSB'
    allfiles = glob.glob(folder_path + "/*.csv")
    final_data = numpy.zeros((1, col_idx2-col_idx1))

    for file_ in sorted(allfiles):
        df = pandas.read_csv(file_, delimiter=',', skiprows=26, header=None)
        df = df[:-1]
        df = df.apply(pandas.to_numeric, args=('coerce',))
        data_arr = df.as_matrix(columns=df.columns[col_idx1:col_idx2])

        #data_arr = data_arr[0:-2, :]
        final_data = numpy.concatenate((final_data, data_arr), axis=0)

    final_data = numpy.delete(final_data, (0), axis=0)
    return final_data


def read_weather_files():
    folder_path = r'/home/sseslab/Documents/SLC PSB data/WBB Weather Data/WBB'
    allfiles = glob.glob(folder_path + "/*.csv")
    #final_data = numpy.zeros((1, col_idx2-col_idx1))
    final_data = pandas.DataFrame()

    for file_ in sorted(allfiles):
        df = pandas.read_csv(file_, delimiter=',', skiprows=7)
        #df = df[:-1]
        #df = df.apply(pandas.to_numeric, args=('coerce',))
        #data_arr = df.as_matrix(columns=df.columns[col_idx1:col_idx2])

        #data_arr = data_arr[0:-2, :]
        #final_data = numpy.concatenate((final_data, data_arr), axis=0)
        final_data = final_data.append(df)

    #final_data = numpy.delete(final_data, (0), axis=0) #deleting first 0 row

    final_data = final_data[:-1] #deleting last row

    return final_data


def split_month(data_arr, month_num, data_interval, offset):
    day_list = numpy.array([31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31])

    if month_num==1:
        z1 = 0
    else:
        z1 = numpy.sum(day_list[0:(month_num-1)])

    print z1
    n1 = z1*(24*(60/data_interval)) + offset
    print n1
    n2 = n1 + (day_list[month_num-1])*(24*(60/data_interval))
    print n2

    data_final = data_arr[n1:n2, :]

    return data_final



def interpolate_nans(X):
    """Overwrite NaNs with column value interpolations."""
    for j in range(X.shape[1]):
        mask_j = numpy.isnan(X[:,j])
        X[mask_j,j] = numpy.interp(numpy.flatnonzero(mask_j), numpy.flatnonzero(~mask_j), X[~mask_j, j])
    return X


def fix_data(X):

    X_new = numpy.zeros((len(X), 1))

    for row in range(1, len(X)):
        X_new[row] = X[row] - X[row-1]

        if X_new[row] < 0:
            X_new[row] = 0

    X_new[0] = 0

    return X_new

def PSB_elec_split(X):
    conv_critical = X[:, 0]
    crac_critical = X[:, 1]
    crac_normal = X[:, 2]
    conv_normal = X[:, 3]
    HVAC_critical = X[:, 4]
    Elev = X[:, 5]
    HVAC_normal = X[:, 6]
    elec_total = X[:, 14]

    return conv_critical, crac_critical, crac_normal, conv_normal, HVAC_critical, Elev, HVAC_normal, elec_total


def give_time(date_string):
    format = '%m/%d/%y %I:%M %p'
    my_date = datetime.strptime(date_string, format)


    return my_date

def give_time_weather(date_string):
    format = '%m/%d/%Y %H:%M'
    my_date = datetime.strptime(date_string, format)


    return my_date


def PSB_time_array(date_array):

    delta_arr = numpy.zeros(len(date_array),)
    delta = numpy.zeros(len(date_array), )
    ref_date = date_array[0]
    print ref_date
    count = 0


    for date_string in date_array:
        a = give_time(date_string) -give_time(ref_date)
        delta_temp = a.total_seconds()/60
        delta_arr[count] = int(delta_temp)

        count += 1



    for i in range(1, len(delta_arr)):
        delta[i] = delta_arr[i] - delta_arr[i-1]

    count = 1
    count_v = []
    count_v.append(0)

    while count < len(delta):
        if delta[count] < 5:
            sum_val = delta[count]
            del_idx = count

            while sum_val < 5:
                del_idx += 1
                sum_val = sum_val + delta[del_idx]

            count = del_idx


        count_v.append(count)
        count += 1


    new_date_array = date_array[count_v]

    return delta, new_date_array

# SELECT DATE FUNCTION FOR PSB

def PSB_select_date(date_array, date_start, date_end):
    #date_start = give_time('6/01/15 12:00 AM')
    #date_end = give_time('6/30/15 11:59 PM')
    date_start = give_time(date_start)
    date_end = give_time(date_end)
    idx_list = []
    count = 0

    for date_string in date_array:
        date_temp = give_time(date_string) - date_start
        date_temp1 = date_temp.total_seconds()

        date_temp = date_end - give_time(date_string)
        date_temp2 = date_temp.total_seconds()


        if date_temp1 >= 0 and date_temp2 >= 0:

            idx_list += [count]

        count += 1

    start_idx = idx_list[0]
    end_idx = idx_list[-1]

    return start_idx, end_idx

# Select Date for weather

def weather_select_date(date_array, date_start, date_end):

    date_start = give_time(date_start)
    date_end = give_time(date_end)
    idx_list = []
    count = 0

    for date_string in date_array:

        date_string = str(date_string)

        #if date_string.endswith('MDT'):
            #date_string = date_string[:-4]


        date_temp = give_time_weather(date_string) - date_start
        date_temp1 = date_temp.total_seconds()
        #print date_string

        date_temp = date_end - give_time_weather(date_string)
        date_temp2 = date_temp.total_seconds()
        #print date_temp


        if date_temp1 >= 0 and date_temp2 >= 0:

            idx_list += [count]

        count += 1

    #print idx_list
    start_idx = idx_list[0]
    end_idx = idx_list[-1]

    return start_idx, end_idx



def find_timesteps(start_date, end_date):
    delta = give_time(end_date) - give_time(start_date)
    delta_t = delta.total_seconds()/60
    timestep = delta_t/5

    return timestep


def find_hourly_timesteps(start_date, end_date):
    delta = give_time(end_date) - give_time(start_date)
    delta_t = delta.total_seconds()/3600
    timestep = delta_t

    return timestep

def convert_to_timetep(delta):

    time_array = numpy.zeros(len(delta),)

    for i in range(1, len(delta)):
        time_array[i] = time_array[i-1] + delta[i]

    timestep_array = time_array

    for i in range(0, len(delta)):
        timestep_array[i] = int(round(time_array[i])/5)

    return timestep_array



def convert_to_hourly_timestep(delta):

    time_array = numpy.zeros(len(delta),)

    for i in range(1, len(delta)):
        time_array[i] = time_array[i-1] + delta[i]

    timestep_array = time_array

    for i in range(0, len(delta)):
        timestep_array[i] = int(round(time_array[i])/60)

    return timestep_array



def arrange_array(X, timestep_array, timestep_total):
    timestep_total = int(round(timestep_total))

    row_max, col_max = X.shape


    X_new = numpy.zeros((timestep_total, col_max))

    #print "timestep_array"
    #print timestep_array

    for i in range(0, len(timestep_array)):
        X_new[timestep_array[i], :] = X[i, :]

    # Make NaN whenever the list is not in array

    for i in range(0, len(X_new)):
        if i not in timestep_array:
            X_new[i, :] = numpy.nan



    last_count = int(round(timestep_array[-1]))

    #print last_count

    #if last_count < timestep_total:
        #for j in range(last_count+1, timestep_total):
           # X_new[j, :] = X[-1, :]

    return X_new


def compile_features(X, date_start, real_res):


    row_max, col_max = X.shape


    max_time = (24*60)/(real_res)
    #print max_time

    t24 = numpy.arange(1.0, max_time+1)
    total_day = row_max/max_time



    time_array = numpy.tile(t24, total_day)

    for i in range(0, len(time_array)):
        time_array[i] = float(time_array[i])/max_time

    #time_temp = time_array[0]
    #time_array = numpy.append(time_array, time_temp)

    my_date = give_time(date_start)
    current_date = my_date
    day_init = my_date.weekday()

    day_mat = numpy.zeros((row_max, 7))
    month_mat = numpy.zeros((row_max, 1))
    daycount_mat = numpy.zeros((row_max, 1))

    for i in range(0, total_day):
        res_val = (i + day_init) % 7

       # if res_val > 4:
        day_mat[i*max_time: (i+1)*max_time, res_val] = 1
        month_mat[i*max_time: (i+1)*max_time] = current_date.month
        daycount_mat[i*max_time: (i+1)*max_time] = current_date.day
        current_date += timedelta(days=1)



    day_mat[-1, :] = day_mat[-2, :]


    X_sch = numpy.concatenate((time_array[:, None], day_mat, daycount_mat, month_mat), axis=1)

    return X_sch




def compile_features_101(X, date_start, real_res):


    row_max, col_max = X.shape


    max_time = (24*60)/(real_res)
    print max_time

    t24 = numpy.arange(1.0, max_time+1)
    total_day = row_max/max_time



    time_array = numpy.tile(t24, total_day)

    for i in range(0, len(time_array)):
        time_array[i] = float(time_array[i])/max_time

    #time_temp = time_array[0]
    #time_array = numpy.append(time_array, time_temp)

    my_date = give_time(date_start)
    current_date = my_date
    day_init = my_date.weekday()

    day_mat = numpy.zeros((row_max, 7))
    month_mat = numpy.zeros((row_max, 1))
    daycount_mat = numpy.zeros((row_max, 1))

    for i in range(0, total_day):
        res_val = (i + day_init) % 7

       # if res_val > 4:
        day_mat[i*max_time: (i+1)*max_time, res_val] = 1
        month_mat[i*max_time: (i+1)*max_time] = current_date.month
        daycount_mat[i*max_time: (i+1)*max_time] = current_date.day
        current_date += timedelta(days=1)



    day_mat[-1, :] = day_mat[-2, :]

    #print time_array
    #print day_mat.shape
    #day_mat removed
    X_sch = numpy.concatenate((time_array[:, None], daycount_mat, month_mat), axis=1)

    return X_sch





#Import weather functions---------------------------------
def read_weather_csv(df, date_start, date_end):
    #file_ = r'/home/sseslab/Documents/SLC PSB data/SLC PSB/PSB_Meters4.csv'

    #df = pandas.read_csv(file_, delimiter=',', skiprows=7)

    datetime_pandas = df.Date_Time
    datetime_mat = datetime_pandas.as_matrix()


    count = 0

    for date_string in datetime_mat:

        date_string = str(date_string)

        if date_string.endswith('MDT') or date_string.endswith('MST'):
            date_string = date_string[:-4]
            datetime_mat[count] = date_string
            count += 1





    start_idx, end_idx = weather_select_date(datetime_mat, date_start, date_end)
    datetime_select = datetime_mat[start_idx:end_idx + 1]

    print datetime_select
    delta, datetime_new = weather_time_array(datetime_select)

    time_array = convert_to_timetep(delta)



    timestep_total = find_timesteps(date_start, date_end)


    df = df.apply(pandas.to_numeric, args=('coerce',))
    data_arr = df.as_matrix(columns=df.columns[3:6])
    data_arr = data_arr[start_idx:end_idx+1, :]
    data_arr = interpolate_nans(data_arr)
    data_arr = arrange_array(data_arr, time_array, timestep_total)
    weather_init = data_arr



    data_arr = df.as_matrix(columns=df.columns[10:11])
    data_arr = data_arr[start_idx:end_idx + 1, :]
    data_arr = interpolate_nans(data_arr)
    data_arr = arrange_array(data_arr, time_array, timestep_total)
    solar_rad = data_arr

    weather_out = numpy.concatenate((weather_init, solar_rad), axis=1)


    return weather_out




def give_weather_time(date_string):
    new_date = parser.parse(date_string)
    new_date = str(new_date)

    format = '%Y-%m-%d %H:%M:%S'
    #print new_date
    new_date = datetime.strptime(new_date, format)

    return new_date


def weather_time_array(date_array):
    delta_arr = numpy.zeros(len(date_array),)
    delta = numpy.zeros(len(date_array), )
    ref_date = date_array[0]

    count = 0

    print date_array.shape
    print date_array

    for date_string in date_array:

        #if date_string.endswith('MDT'):
            #date_string = date_string[:-4]

        #print date_string
        a = give_weather_time(date_string) -give_weather_time(ref_date)
        delta_temp = a.total_seconds()/60
        delta_arr[count] = int(delta_temp)

        count += 1



    for i in range(1, len(delta_arr)):
        delta[i] = delta_arr[i] - delta_arr[i-1]

    count = 1
    count_v = []
    count_v.append(0)

    while count < len(delta):
        if delta[count] < 5:
            sum_val = delta[count]
            del_idx = count

            while sum_val < 5:
                del_idx += 1
                sum_val = sum_val + delta[del_idx]

            count = del_idx

        if delta[count] > 5:
            print count, delta[count]
            #print date_array[count]

        count_v.append(count)
        count += 1


    new_date_array = date_array[count_v]
    return delta, new_date_array


####These functions are for fixing interval

def fix_energy_intervals(X, min_res, real_res):

    skip_factor = int(round(real_res / min_res))
    row_max, col_max = X.shape
    row_new = int(round((row_max - 1)/skip_factor)) + 1
    X_new = numpy.zeros((row_new, col_max))

    count = 0

    for i in range(0, row_new):
        X_new[i] = numpy.sum(X[count:count+skip_factor])
        count = count + skip_factor

    return X_new


def fix_weather_intervals(X, min_res, real_res):

    skip_factor = int(round(real_res / min_res))
    row_max, col_max = X.shape
    row_new = int(round((row_max - 1)/skip_factor)) + 1
    X_new = numpy.zeros((row_new, col_max))

    count = 0

    for i in range(0, row_new):
        for j in range(0, col_max):
            X_new[i, j] = numpy.sum(X[count:count+skip_factor, j])/skip_factor
        count = count + skip_factor

    return X_new


def get_normalize_params(X):

    max_array = numpy.nanmax(X, axis=0)
    min_array = numpy.nanmin(X, axis=0)


    return min_array, max_array

def aggregate_data(X, conv_factor):
    row_max, col_max = X.shape
    #conv_factor = 24
    day_max = int(numpy.trunc(row_max/conv_factor))

    min_X = numpy.zeros((day_max, col_max))
    max_X = numpy.zeros((day_max, col_max))
    sum_X = numpy.zeros((day_max, col_max))
    mean_X = numpy.zeros((day_max, col_max))

    for i in range(0, day_max):
        min_X[i, :] = numpy.amin(X[i*conv_factor:(i+1)*conv_factor, :], axis=0)
        max_X[i, :] = numpy.amax(X[i*conv_factor:(i+1)*conv_factor, :], axis=0)
        sum_X[i, :] = numpy.sum(X[i*conv_factor:(i+1)*conv_factor, :], axis=0)
        mean_X[i, :] = sum_X[i, :]/conv_factor

    return mean_X, sum_X, min_X, max_X


def fix_high_points(X):

    for row in range(0, len(X)):
        if X[row] > 3*numpy.nanmean(X):
            X[row] = numpy.nanmean(X)


    return X

def find_error(H_t, H_e, Y_lstm):
    H_t = numpy.reshape(H_t, (H_t.shape[0] * 24, 1))
    H_e = numpy.reshape(H_e, (H_e.shape[0] * 24, 1))
    Y_lstm = numpy.reshape(Y_lstm, (Y_lstm.shape[0] * 24, 1))

    e_deep = (MathFunctions.rms_flat(Y_lstm - H_e)) / (MathFunctions.rms_flat(H_e))
    e_deep2 = (MathFunctions.rms_flat(Y_lstm - H_e)) / (MathFunctions.rms_flat(H_t))

    return e_deep, e_deep2


def find_error2(H_t, H_e, Y_lstm, tsteps):
    H_t = numpy.reshape(H_t, (H_t.shape[0] *tsteps, 1))
    H_e = numpy.reshape(H_e, (H_e.shape[0] * tsteps, 1))
    Y_lstm = numpy.reshape(Y_lstm, (Y_lstm.shape[0] * tsteps, 1))

    e_deep = (MathFunctions.rms_flat(Y_lstm - H_e)) / (MathFunctions.rms_flat(H_e))
    e_deep2 = (MathFunctions.rms_flat(Y_lstm - H_e)) / (MathFunctions.rms_flat(H_t))

    return e_deep, e_deep2



def calculate_entropy(X):

    X = X[~numpy.isnan(X)]
    unique, counts = numpy.unique(X, return_counts=True)

    counts = counts.astype(float)
    #counts = counts[(counts > 10)]

    pk = ((counts)/(numpy.sum(counts)))
    idx = numpy.where(pk > 0.01)
    #pk = pk[(pk > 0.01)]
    unique = unique[idx]

    S = scipy.stats.entropy(pk)

    return S, unique, pk

def convert_to_discrete(X, discrete_values):
    X_new = X

    for i in range(0, len(X)):
        temp_val = abs(discrete_values - X[i])

        min_idx = numpy.argmin(temp_val)
        X_new[i] = discrete_values[min_idx]

    return X_new

def make_binary(X):
    X_new = X

    for i in range(0, X.shape[0]):
        for j in range(0, X.shape[1]):
            temp = X[i, j, :]
            idx = numpy.argmax(temp)

            for k in range(0, X.shape[2]):
                if k == idx:
                    X_new[i, j, k] = 1
                else:
                    X_new[i, j, k] = 0


    return X_new

def convert_to_continous_3D(X, unique):

    print X.shape
    X_new  = numpy.zeros((X.shape[0], X.shape[1], 1))

    for i in range(0, X.shape[0]):
        for j in range(0, X.shape[1]):
            temp = X[i, j, :]
            #print temp
            idx = numpy.where(temp == 1)
            #print idx
            idx = idx[0]
            #print idx
            X_new[i, j, 0] = unique[idx]

    return X_new

def one_hot_encoder_101(X):
    encoder = LabelEncoder()
    encoder.fit(X)
    encoded_Y = encoder.transform(X)
    dummy_y = np_utils.to_categorical(encoded_Y)

    return dummy_y

def fix_discrete(X, unique):

    X_new = X

    for i in range(0, len(X)):
        diff = abs(unique - X[i])

        idx = numpy.argmin(diff)
        X_new[i] = unique[idx]

    return X_new


def normalize_2D(X):
    X_new = (X - X.mean(axis=0)) / X.std(axis=0)
    nan_idx = numpy.isnan(X_new)
    X_new[nan_idx] = 0

    return X_new





def read_dataport_csv(file_):

    df = pandas.read_csv(file_, delimiter=',')
    df = df.apply(pandas.to_numeric, args=('coerce',))
    data_arr = df.as_matrix(columns=df.columns[2:3])

    return data_arr


###Function to read weather_austin_file

def read_weather_austin(file_, date_start, date_end):

    df = pandas.read_csv(file_, delimiter=',', skiprows=7)

    datetime_pandas = df.Date_Time
    datetime_mat = datetime_pandas.as_matrix()

    count = 0

    for date_string in datetime_mat:

        date_string = str(date_string)

        if date_string.endswith('CDT') or date_string.endswith('CST'):
            date_string = date_string[:-4]
            datetime_mat[count] = date_string
            count += 1

    start_idx, end_idx = weather_select_date(datetime_mat, date_start, date_end)
    datetime_select = datetime_mat[start_idx:end_idx + 1]


    delta, datetime_new = weather_time_array(datetime_select)
    print delta

    time_array = convert_to_hourly_timestep(delta)
    print time_array
    timestep_total = int(find_hourly_timesteps(date_start, date_end))
    print timestep_total

    df = df.apply(pandas.to_numeric, args=('coerce',))
    data_arr = df.as_matrix(columns=df.columns[2:7])
    data_arr = data_arr[start_idx:end_idx + 1, :]
    data_arr = interpolate_nans(data_arr)
    data_arr = arrange_array(data_arr, time_array, timestep_total)
    weather_out = data_arr



    return df, weather_out


def find_val_set(init_num, X, Y):

    month_array = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    #rand_array = numpy.zeros(len(month_array),)

    X_val = numpy.zeros((1, X.shape[1]))
    Y_val = numpy.zeros((1, Y.shape[1]))

    for i in range(0, 12):
        rand_num = random.randint(0, month_array[i])
        rand1 = init_num + rand_num*24
        rand2 = init_num + (rand_num+1)*24

        X_temp = X[rand1:rand2, :]
        Y_temp = Y[rand1:rand2, :]
        X_val = numpy.append(X_val, X_temp, axis=0)
        Y_val = numpy.append(Y_val, Y_temp, axis=0)

    X_val = X_val[1:, :]
    Y_val = Y_val[1:, :]
    return X_val, Y_val


def find_val_set2(init_num, X, Y, X2, Y2, n):

    #n is a multiple of 24, where n denotes the number of timesteps in an hour
    month_array = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
    #rand_array = numpy.zeros(len(month_array),)

    X_val = numpy.zeros((1, X.shape[1]))
    Y_val = numpy.zeros((1, Y.shape[1]))
    X_val2 = numpy.zeros((1, X2.shape[1]))
    Y_val2 = numpy.zeros((1, Y2.shape[1]))

    for i in range(0, 12):
        rand_num = random.randint(0, month_array[i])
        rand1 = init_num + rand_num*24
        rand2 = init_num + (rand_num+1)*24
        rand3 = init_num + rand_num*24*n
        rand4 = init_num + (rand_num+1)*24*n

        X_temp = X[rand1:rand2, :]
        Y_temp = Y[rand1:rand2, :]
        X_val = numpy.append(X_val, X_temp, axis=0)
        Y_val = numpy.append(Y_val, Y_temp, axis=0)

        X_temp = X2[rand3:rand4, :]
        Y_temp = Y2[rand3:rand4, :]
        X_val2 = numpy.append(X_val2, X_temp, axis=0)
        Y_val2 = numpy.append(Y_val2, Y_temp, axis=0)


    X_val = X_val[1:, :]
    Y_val = Y_val[1:, :]
    X_val2 = X_val2[1:, :]
    Y_val2 = Y_val2[1:, :]

    return X_val, Y_val, X_val2, Y_val2


def preprocess_energyData_101(file_):
    data = read_dataport_csv(file_)
    H_t = data[0:len(data) - 1]  # removing las tentry

    # Normalizing H_t
    H_min, H_max = get_normalize_params(H_t)
    #H_t = H_t / H_max

    return H_t, H_max

def preprocess_weatherData_101(file_, date_start, date_end):
    df, weather_data = read_weather_austin(file_, date_start, date_end)
    weather_data = interpolate_nans(weather_data)

    # Appending the initial weather file
    w0 = weather_data[0, :]
    w0 = w0[None, :]
    weather = numpy.append(w0, weather_data, axis=0)

    return weather


############Datafunctionsw for multi-timescale analysis
def get_weather_1min(X_hour, std_intv1, std_intv2):

    factor = std_intv1/std_intv2
    total_hour = len(X_hour)
    total_min = total_hour*factor

    #Constructing the minute vector
    min_v = (numpy.arange(1, factor+1))/float(std_intv1)
    t_min = numpy.tile(min_v, total_hour)
    t_min = t_min[:, None]
    X_out = numpy.zeros((total_min, X_hour.shape[1]))

    for i in range(0, total_hour):
        X_out[i*factor:(i+1)*factor, :] = X_hour[i, :]


    X_out = numpy.concatenate((X_out, t_min), axis=1)

    return X_out


def reduce_by_sum(Y, tsteps):
    num = len(Y)/tsteps
    Y_new = numpy.zeros((num, Y.shape[1]))

    for i in range(0, num):
        Y_new[i, :] = numpy.sum(Y[i*tsteps:(i+1)*tsteps, :])

    return Y_new